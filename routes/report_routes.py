from flask import (
    Blueprint,
    render_template,
    request,
    session,
    redirect,
    url_for,
    send_file,
    flash,
)

from utils.decorators import permission_required
from models.role import Role
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database.database import db
from models.bdc_request import BDCRequest

report_bp = Blueprint(
    "reports",
    __name__
)


# =====================================================
# Shared Filter Builder
# =====================================================

def build_filtered_query():
    """
    Build a filtered BDCRequest query.

    This helper is reused by:
        - Reports
        - Excel Export
        - CSV Export
        - Print Report
    """

    status = request.args.get("status", "")
    bdc_agent = request.args.get("bdc_agent", "")
    sales_rep = request.args.get("sales_rep", "")
    request_type = request.args.get("request_type", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    query = BDCRequest.query

    if status:
        query = query.filter(BDCRequest.status == status)

    if bdc_agent:
        query = query.filter(BDCRequest.bdc_agent == bdc_agent)

    if sales_rep:
        query = query.filter(BDCRequest.sales_rep == sales_rep)

    if request_type:
        query = query.filter(BDCRequest.request_type == request_type)

    if date_from:
        start_date = datetime.strptime(date_from, "%Y-%m-%d")
        query = query.filter(BDCRequest.created_at >= start_date)

    if date_to:
        end_date = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
        query = query.filter(BDCRequest.created_at < end_date)

    filters = {
        "status": status,
        "bdc_agent": bdc_agent,
        "sales_rep": sales_rep,
        "request_type": request_type,
        "date_from": date_from,
        "date_to": date_to,
    }

    return query, filters


# =====================================================
# Reports
# =====================================================

@report_bp.route("/reports")
@permission_required("reports")
def reports():

    if "user_id" not in session:
        return redirect(url_for("login"))


    # ---------------------------------------------
    # Filtered Query
    # ---------------------------------------------

    query, filters = build_filtered_query()

    requests = (
        query
        .order_by(BDCRequest.created_at.desc())
        .all()
    )

    # ============================================
    # Summary Counts
    # ============================================

    total_requests = len(requests)

    in_progress_count = sum(
        1 for request in requests
        if request.status == "In Progress"
    )

    completed_count = sum(
        1 for request in requests
        if request.status == "Completed"
    )

    cancelled_count = sum(
        1 for request in requests
        if request.status == "Cancelled"
    )

    total_requests = len(requests)

    # ---------------------------------------------
    # Dropdown Values
    # ---------------------------------------------

    bdc_agents = (
        db.session.query(BDCRequest.bdc_agent)
        .filter(BDCRequest.bdc_agent.isnot(None))
        .filter(BDCRequest.bdc_agent != "")
        .distinct()
        .order_by(BDCRequest.bdc_agent)
        .all()
    )

    sales_reps = (
        db.session.query(BDCRequest.sales_rep)
        .filter(BDCRequest.sales_rep.isnot(None))
        .filter(BDCRequest.sales_rep != "")
        .distinct()
        .order_by(BDCRequest.sales_rep)
        .all()
    )

    request_types = (
        db.session.query(BDCRequest.request_type)
        .filter(BDCRequest.request_type.isnot(None))
        .filter(BDCRequest.request_type != "")
        .distinct()
        .order_by(BDCRequest.request_type)
        .all()
    )

    # ---------------------------------------------
    # Render Page
    # ---------------------------------------------

    return render_template(
        "reports.html",
        username=session["username"],
        role=session["role"],

        total_requests=total_requests,
        in_progress_count=in_progress_count,
        completed_count=completed_count,
        cancelled_count=cancelled_count,

        requests=requests,

        status=filters["status"],
        bdc_agent=filters["bdc_agent"],
        sales_rep=filters["sales_rep"],
        request_type=filters["request_type"],
        date_from=filters["date_from"],
        date_to=filters["date_to"],

        bdc_agents=bdc_agents,
        sales_reps=sales_reps,
        request_types=request_types,
    )

# =====================================================
# Export Excel
# =====================================================

@report_bp.route("/reports/export/excel")
def export_excel():

    if "user_id" not in session:
        return redirect(url_for("login"))

    query, filters = build_filtered_query()

    requests = (
        query
        .order_by(BDCRequest.created_at.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "BDC Report"

    ws.merge_cells("A1:I1")
    ws.merge_cells("A2:I2")

    ws["A1"] = "Honda of Cleveland Heights"
    ws["A2"] = "Business Development Center Report"

    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"

    ws["A1"].font = Font(size=18, bold=True)
    ws["A2"].font = Font(size=14, bold=True)
    ws["A3"].font = Font(italic=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Ticket #",
        "Date Created",
        "Customer",
        "Phone",
        "Request Type",
        "Sales Representative",
        "Sales Manager",
        "Status",
        "BDC Agent",
    ]

    header_row = 5

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)

        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

        row = 6

        for request in requests:

            customer_name = (
                f"{request.first_name} "
                f"{request.middle_name + ' ' if request.middle_name else ''}"
                f"{request.last_name}"
            ).strip()

            ws.cell(row=row, column=1).value = request.ticket_number
            ws.cell(row=row, column=2).value = request.created_at.strftime("%Y-%m-%d")
            ws.cell(row=row, column=3).value = customer_name
            ws.cell(row=row, column=4).value = request.phone
            ws.cell(row=row, column=5).value = request.request_type
            ws.cell(row=row, column=6).value = request.sales_rep
            ws.cell(row=row, column=7).value = request.sales_manager
            ws.cell(row=row, column=8).value = request.status
            ws.cell(row=row, column=9).value = request.bdc_agent

            ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")

            if request.status == "Completed":
                ws.cell(row=row, column=8).fill = PatternFill(
                    fill_type="solid",
                    fgColor="C6EFCE"
                )

            elif request.status == "In Progress":
                ws.cell(row=row, column=8).fill = PatternFill(
                    fill_type="solid",
                    fgColor="FFF2CC"
                )

            elif request.status == "Cancelled":
                ws.cell(row=row, column=8).fill = PatternFill(
                    fill_type="solid",
                    fgColor="F4CCCC"
                )

            for col in range(1, 10):
                ws.cell(row=row, column=col).border = thin_border

            row += 1

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:I{row - 1}"

    # Auto-size columns
    for column_cells in ws.columns:
        max_length = 0

        for cell in column_cells:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = max_length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="BDC_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ============================================
# Role Access Management
# ============================================

@report_bp.route(
    "/settings/role-permissions",
    methods=["GET", "POST"]
)
@report_bp.route(
    "/settings/role-permissions/<int:role_id>",
    methods=["GET", "POST"]
)
@permission_required("settings")
def role_access_management(role_id=None):

    if "user_id" not in session:
        return redirect(url_for("login"))

    roles = Role.query.order_by(Role.name).all()

    selected_role = None
    permissions = None

    if role_id:
        selected_role = Role.query.get_or_404(role_id)
        permissions = selected_role.permissions

    if request.method == "POST" and permissions:

        permissions.dashboard = "dashboard" in request.form
        permissions.new_bdc_request = "new_bdc_request" in request.form
        permissions.bdc_request_queue = "bdc_request_queue" in request.form
        permissions.delete_bdc_request = "delete_bdc_request" in request.form
        permissions.employee_management = "employee_management" in request.form
        permissions.reports = "reports" in request.form
        permissions.settings = "settings" in request.form

        db.session.commit()

        flash("Permissions updated successfully.", "success")

        return redirect(
            url_for(
                "reports.role_access_management",
                role_id=selected_role.id
            )
        )

    return render_template(
        "settings/role_permissions.html",
        roles=roles,
        selected_role=selected_role,
        permissions=permissions
    )